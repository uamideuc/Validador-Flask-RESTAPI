"""
Exportador de Excel de validación - Versión 2.0 con coloreo de celdas y columnas val_*
"""
import pandas as pd
import os
import tempfile
import json
from typing import Dict, Any, List, Tuple
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from ....core.models import VariableCategorization
from ....core.services.file_handling.file_parser import FileParser


# Colores de categorías (idénticos al frontend)
CATEGORY_COLORS = {
    'instrument_vars': 'FF1976D2',      # Azul
    'item_id_vars': 'FF388E3C',         # Verde
    'metadata_vars': 'FFF57C00',        # Naranja
    'classification_vars': 'FF7B1FA2'   # Púrpura
}


class ValidationExcelExporter:
    """Exportador específico para Excel de validación con celdas coloreadas y columnas val_*"""

    def __init__(self, session_id: str):
        self.session_id = session_id
        self.db = self._get_db_manager()

    def export(self, validation_session_id: int) -> Dict[str, Any]:
        """
        Exportar Excel de validación con:
        1. Celdas problemáticas coloreadas según categoría
        2. Columnas val_* en lugar de ESTADO_VALIDACION
        """
        try:
            # Get validation session data
            validation_session = self.db.get_validation_session(validation_session_id)
            if not validation_session:
                return {
                    'success': False,
                    'error': 'Sesión de validación no encontrada'
                }

            # Load original data and validation results
            validation_results = validation_session['validation_results']
            if isinstance(validation_results, str):
                validation_results = json.loads(validation_results)

            # Load original data from file using FileParser (handles CSV with any delimiter)
            original_file_path = validation_session['file_path']
            sheet_name = validation_session.get('sheet_name')  # Get sheet_name for Excel files
            parser = FileParser()
            original_data = parser.parse_file(original_file_path, sheet_name)

            categorization_dict = validation_session['categorization']
            if isinstance(categorization_dict, str):
                categorization_dict = json.loads(categorization_dict)
            categorization = VariableCategorization(**categorization_dict)

            # Build problems map for each cell
            cell_problems = self._build_cell_problems_map(original_data, categorization)

            # Create annotated data with val_* columns
            annotated_data = self._create_annotated_data_with_val_columns(
                original_data,
                categorization,
                cell_problems
            )

            # Create Excel file
            filename = f"validation_excel_{validation_session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            temp_dir = tempfile.gettempdir()
            file_path = os.path.join(temp_dir, filename)

            # Write data to Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet 1: Datos originales con val_* columns
                annotated_data.to_excel(writer, sheet_name='Datos_Validacion', index=False)

                # Sheet 2: Resumen validación (mejorado)
                validation_summary = self._create_validation_summary_sheet(
                    validation_results,
                    categorization,
                    cell_problems
                )
                validation_summary.to_excel(writer, sheet_name='Resumen_Validacion', index=False)

            # Apply cell and column formatting with openpyxl
            self._apply_excel_formatting(
                file_path,
                annotated_data,
                categorization,
                cell_problems
            )

            # Registrar en database
            export_id = self.db.create_export_record(
                validation_session_id=validation_session_id,
                session_id=self.session_id,
                export_type='validation_excel',
                file_path=file_path
            )

            return {
                'success': True,
                'export_id': export_id,
                'filename': filename,
                'file_path': file_path,
                'export_type': 'validation_excel'
            }

        except Exception as e:
            import traceback
            return {
                'success': False,
                'error': f'Error en exportación Excel: {str(e)}',
                'traceback': traceback.format_exc()
            }

    def _build_cell_problems_map(
        self,
        data: pd.DataFrame,
        categorization: VariableCategorization
    ) -> Dict[str, Dict[int, str]]:
        """
        Construir mapa de problemas por celda
        Retorna: {'variable_name': {row_index: 'TIPO_PROBLEMA', ...}, ...}
        """
        problems = {}

        # 1. Detectar duplicados en item_id_vars
        for var in categorization.item_id_vars:
            if var not in data.columns:
                continue

            problems[var] = {}

            # Agrupar por instrument_vars si existen
            if categorization.instrument_vars:
                instruments = self._group_by_instruments(data, categorization)
                for instrument_key, instrument_data in instruments.items():
                    if var in instrument_data.columns:
                        # Encontrar valores duplicados
                        non_null = instrument_data[var].dropna()
                        value_counts = non_null.value_counts()
                        duplicated_values = value_counts[value_counts > 1].index

                        # Marcar todas las filas con valores duplicados
                        for dup_val in duplicated_values:
                            mask = instrument_data[var] == dup_val
                            dup_indices = instrument_data[mask].index.tolist()
                            for idx in dup_indices:
                                problems[var][idx] = 'DUPLICADO'
            else:
                # Sin instrumentos, buscar duplicados en toda la data
                non_null = data[var].dropna()
                value_counts = non_null.value_counts()
                duplicated_values = value_counts[value_counts > 1].index

                for dup_val in duplicated_values:
                    mask = data[var] == dup_val
                    dup_indices = data[mask].index.tolist()
                    for idx in dup_indices:
                        problems[var][idx] = 'DUPLICADO'

            # Detectar valores faltantes en item_id_vars (igual que metadata_vars)
            missing_mask = data[var].isna()
            missing_indices = data[missing_mask].index.tolist()

            for idx in missing_indices:
                # Si ya tiene un problema (ej: duplicado), combinar
                if idx in problems[var]:
                    problems[var][idx] = f"{problems[var][idx]}+VALOR_FALTANTE"
                else:
                    problems[var][idx] = 'VALOR_FALTANTE'

        # 2. Detectar valores faltantes en metadata_vars
        for var in categorization.metadata_vars:
            if var not in data.columns:
                continue

            if var not in problems:
                problems[var] = {}

            # Encontrar valores faltantes
            missing_mask = data[var].isna()
            missing_indices = data[missing_mask].index.tolist()

            for idx in missing_indices:
                # Si ya tiene un problema, combinar
                if idx in problems[var]:
                    problems[var][idx] = f"{problems[var][idx]}+VALOR_FALTANTE"
                else:
                    problems[var][idx] = 'VALOR_FALTANTE'

        # 3. Detectar valores faltantes en instrument_vars (crítico)
        for var in categorization.instrument_vars:
            if var not in data.columns:
                continue

            if var not in problems:
                problems[var] = {}

            missing_mask = data[var].isna()
            missing_indices = data[missing_mask].index.tolist()

            for idx in missing_indices:
                if idx in problems[var]:
                    problems[var][idx] = f"{problems[var][idx]}+VALOR_FALTANTE"
                else:
                    problems[var][idx] = 'VALOR_FALTANTE'

        return problems

    def _create_annotated_data_with_val_columns(
        self,
        data: pd.DataFrame,
        categorization: VariableCategorization,
        cell_problems: Dict[str, Dict[int, str]]
    ) -> pd.DataFrame:
        """
        Crear DataFrame con:
        1. TODAS las columnas originales primero (pristinas, en orden exacto)
        2. TODAS las columnas val_* al final (en orden de aparición original)

        Esto permite al usuario:
        - Ver datos originales completos (izquierda)
        - Revisar validaciones (derecha)
        - Corregir datos originales
        - Eliminar todas las val_* de un jalón al final
        """
        # 1. COPIAR TODOS LOS DATOS ORIGINALES (pristinos, mismo orden)
        result = data.copy()

        # 2. Variables que necesitan columnas val_* (solo las categorizadas validadas)
        validated_vars = set(
            categorization.instrument_vars +
            categorization.item_id_vars +
            categorization.metadata_vars +
            categorization.classification_vars
        )

        # 3. AGREGAR TODAS LAS COLUMNAS val_* AL FINAL
        # Recorrer columnas originales en orden para mantener orden de val_*
        for col in data.columns:
            # Solo crear val_* si la variable está categorizada y validada
            if col in validated_vars:
                val_col_name = f'val_{col}'
                val_col_data = []

                # Determinar valor para cada fila
                problems_for_col = cell_problems.get(col, {})
                for idx in data.index:
                    if idx in problems_for_col:
                        val_col_data.append(problems_for_col[idx])
                    else:
                        val_col_data.append('OK')

                # Agregar columna val_* AL FINAL (después de todas las originales)
                result[val_col_name] = val_col_data

        return result

    def _apply_excel_formatting(
        self,
        file_path: str,
        annotated_data: pd.DataFrame,
        categorization: VariableCategorization,
        cell_problems: Dict[str, Dict[int, str]]
    ):
        """
        Aplicar formato Excel:
        1. Colorear celdas problemáticas con color de categoría (fondo)
        2. Colorear texto de columnas val_* según categoría
        """
        wb = load_workbook(file_path)
        ws = wb['Datos_Validacion']

        # Mapeo de variable a categoría
        var_to_category = {}
        for cat_key, vars_list in [
            ('instrument_vars', categorization.instrument_vars),
            ('item_id_vars', categorization.item_id_vars),
            ('metadata_vars', categorization.metadata_vars),
            ('classification_vars', categorization.classification_vars)
        ]:
            for var in vars_list:
                var_to_category[var] = cat_key

        # Obtener mapeo de columna a índice en Excel (1-indexed)
        col_name_to_excel_idx = {}
        for idx, col_name in enumerate(annotated_data.columns, start=1):
            col_name_to_excel_idx[col_name] = idx

        # 1. Colorear celdas problemáticas (datos originales)
        for var, problems in cell_problems.items():
            if var not in col_name_to_excel_idx:
                continue

            excel_col_idx = col_name_to_excel_idx[var]
            category = var_to_category.get(var)

            if not category or category not in CATEGORY_COLORS:
                continue

            color = CATEGORY_COLORS[category]
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

            # Aplicar color a cada celda con problema
            for row_idx in problems.keys():
                # row_idx es 0-indexed del DataFrame, Excel es 1-indexed + header
                excel_row = row_idx + 2  # +1 for header, +1 for 0-indexed
                cell = ws.cell(row=excel_row, column=excel_col_idx)
                cell.fill = fill
                # Texto blanco para mejor contraste
                cell.font = Font(color='FFFFFFFF', bold=True)

        # 2. Colorear texto de columnas val_* SOLO cuando NO es "OK"
        for col_name in annotated_data.columns:
            if not col_name.startswith('val_'):
                continue

            # Extraer nombre de variable original
            original_var = col_name[4:]  # Quitar 'val_'

            if original_var not in var_to_category:
                continue

            category = var_to_category[original_var]
            color = CATEGORY_COLORS[category]

            excel_col_idx = col_name_to_excel_idx[col_name]

            # Aplicar color SOLO a celdas con problemas (no "OK")
            font_with_color = Font(color=color, bold=True)
            font_ok = Font(color='FF808080', bold=False)  # Gris suave para "OK"

            for row_idx in range(2, len(annotated_data) + 2):  # +2 for header and 1-indexed
                cell = ws.cell(row=row_idx, column=excel_col_idx)
                cell_value = str(cell.value).strip()

                # Solo colorear si NO es "OK"
                if cell_value != 'OK':
                    cell.font = font_with_color  # Color de categoría
                else:
                    cell.font = font_ok  # Gris suave

        wb.save(file_path)

    def _group_by_instruments(
        self,
        data: pd.DataFrame,
        categorization: VariableCategorization
    ) -> Dict[str, pd.DataFrame]:
        """Agrupar datos por instrumentos"""
        if not categorization.instrument_vars:
            return {'default': data}

        instruments = {}

        for idx, row in data.iterrows():
            instrument_values = {}
            for var in categorization.instrument_vars:
                if var in data.columns:
                    instrument_values[var] = str(row[var])

            instrument_key = '|'.join([f"{k}:{v}" for k, v in sorted(instrument_values.items())])

            if instrument_key not in instruments:
                instruments[instrument_key] = []

            instruments[instrument_key].append(row.to_dict())

        # Convert to DataFrames
        result = {}
        for key, rows in instruments.items():
            result[key] = pd.DataFrame(rows)

        return result

    def _create_validation_summary_sheet(
        self,
        validation_results: Dict,
        categorization: VariableCategorization,
        cell_problems: Dict[str, Dict[int, str]]
    ) -> pd.DataFrame:
        """Crear sheet de resumen de validación mejorado"""
        summary_info = []

        # Encabezado
        summary = validation_results.get('summary', {})
        summary_info.append(['RESUMEN DE VALIDACIÓN', ''])
        summary_info.append(['', ''])
        summary_info.append(['Fecha de validación', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_info.append(['Total de ítems', summary.get('total_items', 0)])
        summary_info.append(['Total de instrumentos', summary.get('total_instruments', 0)])
        summary_info.append(['Estado general', summary.get('validation_status', 'desconocido').upper()])
        summary_info.append(['', ''])

        # Conteo de problemas por tipo
        total_duplicados = sum(1 for problems in cell_problems.values()
                              for problem in problems.values()
                              if 'DUPLICADO' in problem)
        total_faltantes = sum(1 for problems in cell_problems.values()
                             for problem in problems.values()
                             if 'VALOR_FALTANTE' in problem)
        total_problemas = total_duplicados + total_faltantes

        summary_info.append(['CONTEO DE PROBLEMAS DETECTADOS', ''])
        summary_info.append(['Total de celdas con problemas', total_problemas])
        summary_info.append(['  - Valores duplicados', total_duplicados])
        summary_info.append(['  - Valores faltantes', total_faltantes])
        summary_info.append(['', ''])

        # Detalle por categoría de variables
        summary_info.append(['VARIABLES CATEGORIZADAS', ''])
        summary_info.append(['', ''])

        summary_info.append(['Categoría', 'Variables'])

        if categorization.instrument_vars:
            summary_info.append(['Identificación de Instrumento', ', '.join(categorization.instrument_vars)])

        if categorization.item_id_vars:
            summary_info.append(['Identificación de Ítems', ', '.join(categorization.item_id_vars)])

        if categorization.metadata_vars:
            summary_info.append(['Información Crítica', ', '.join(categorization.metadata_vars)])

        if categorization.classification_vars:
            summary_info.append(['Información Complementaria', ', '.join(categorization.classification_vars)])

        summary_info.append(['', ''])

        # Estado de validaciones
        summary_info.append(['ESTADO DE VALIDACIONES', ''])
        summary_info.append(['', ''])

        duplicate_validation = validation_results.get('duplicate_validation', {})
        summary_info.append(['Validación de Duplicados', 'VÁLIDO ✓' if duplicate_validation.get('is_valid', True) else 'ERRORES ✗'])

        metadata_validation = validation_results.get('metadata_validation', {})
        summary_info.append(['Validación de Metadata', 'VÁLIDO ✓' if metadata_validation.get('is_valid', True) else 'ERRORES ✗'])

        instrument_validation = validation_results.get('instrument_validation', {})
        summary_info.append(['Validación de Instrumentos', 'VÁLIDO ✓' if instrument_validation.get('is_valid', True) else 'ERRORES ✗'])

        summary_info.append(['', ''])
        summary_info.append(['INSTRUCCIONES', ''])
        summary_info.append(['1. Revise las celdas coloreadas en la hoja "Datos_Validacion"', ''])
        summary_info.append(['2. Consulte las columnas val_* para ver el tipo de problema', ''])
        summary_info.append(['3. Corrija los errores en las columnas originales', ''])
        summary_info.append(['4. Elimine todas las columnas val_* antes de usar los datos', ''])

        # Crear DataFrame con nombres de columnas explícitos
        return pd.DataFrame(summary_info, columns=['Métrica', 'Valor'])

    def _get_db_manager(self):
        """Get database manager instance"""
        from flask import current_app
        if current_app.config.get('TESTING'):
            from ....core.database import DatabaseManager
            db_path = current_app.config.get('DATABASE_PATH', 'test.db')
            return DatabaseManager(db_path)
        else:
            from ....core.database import db_manager
            return db_manager
